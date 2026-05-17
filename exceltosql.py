import openpyxl

from openpyxl import load_workbook
wb = load_workbook(filename="dbsyteil3.xlsx")
ws = wb["Ferienwohnung"]
f = open("sqlbefehle.txt", "w")

max_row = ws.max_row
max_col = ws.max_column

for i in range(2, ws.max_row + 1):
	if (ws.cell(row=i, column=1).value != None):
		table = ws.cell(row=i, column=1).value
		f.write(f"CREATE TABLE {table} (\n")
	else:
		# instead of an f loop, use indexes to grab different columns and craft commands
		attribute = ws.cell(row=i, column=2).value
		datatype = ws.cell(row=i, column=4).value
		length = ws.cell(row=i, column=5).value
		m_or_k = ws.cell(row=i, column=6).value
		eindeutig = ws.cell(row=i, column=7).value      
		primarykey = ws.cell(row=i, column=8).value 
		foreignkey = ws.cell(row=i, column=9).value
		
		# initial attribute creation and data type declaration
		match datatype:
			case "INT":
				f.write(f"{attribute}	INT")
			case "String":
				if length != None:
					f.write(f"{attribute}	VARCHAR({length})")
				else: 
					f.write(f"{attribute}	VARCHAR(255)")
			case "Double":
				f.write(f"{attribute}	NUMBER(6,2)")
			case "Date":
				f.write(f"{attribute}	DATE")

		# not null constraint
		if m_or_k == "m":
			f.write("	NOT NULL,\n")
		elif m_or_k == "k":
			f.write(",\n")

		# primary key declaration
		if primarykey == "j":
			f.write(f"CONSTRAINT {table}_pk PRIMARY KEY({attribute})")
		
		# foreign key declaration and skip to the next iteration
		if foreignkey == None:
			f.write(");\n")
			continue
		else:
			f.write(f"CONSTRAINT {table}_fk FOREIGN KEY ({attribute}) REFERENCES {foreignkey});\n")
			continue



		



             
       
           
           

# loop through the rows
# loop through the columns
# if column 1 has a value, create table paranthesis
# if column 2 has a value, create attribute
# if column 4 is string, varchar()
# if column 4 is string and length has value = varchar(value)
# if column 4 is string and legnth is not given, varchar(255)
# if column 6 is m, NOT NULL
# if column 8 is j, primary key
# if column 9 has value, add foreign key

